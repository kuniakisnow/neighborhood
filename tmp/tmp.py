from cProfile import label
import openpyxl
import pprint
import os
import sys


for entry in os.scandir():
    if entry.is_file():
        print(entry.name)
    else:
        print(f'dir:{entry.name}')

wb = openpyxl.load_workbook('./tmp/アンケート結果.xlsx')
print(type(wb))
print(wb.sheetnames)

sheet = wb['Sheet1']
cell = sheet['C1']
print(cell.value)
print(sheet.max_row)

MaxRow = sheet.max_row + 1

#Sheet2を作成
def makeSheet2():
    global sheet
    global wb

    sheet2 = wb['Sheet2']
    sheet2.cell(row=1,column=1,value='項目')
    sheet2.cell(row=1,column=2,value='number')
    sheet2.cell(row=1,column=3,value='班')
    sheet2.cell(row=1,column=4,value='属性')
    sheet2.cell(row=1,column=5,value='結果')

    num = 2
    for i in range(2,MaxRow):
        team = sheet.cell(row=i,column=2).value
    #    print(team)
        for j in range(3,15):
            cell = sheet.cell(row=i,column=j).value
            if cell == 1 : 
                cell = '設問' + str(j - 2)
            if j == 13:
                cell = "" # mask it 
            sheet2.cell(row=num,column=1,value=num)
            sheet2.cell(row=num,column=2,value=i-1)
            sheet2.cell(row=num,column=3,value=team)
            sheet2.cell(row=num,column=4,value=j)
            sheet2.cell(row=num,column=5,value=cell)
            num = num + 1

#全体＋１３チーム　0は全体　14個
NumberTeam = 14
Answer = [[0 for i in range(14)] for j in range(14)]
PercentNumber = [[0 for i in range(14)] for j in range(14)]
Respondents = [0]*14 #回答者数
Label = ["全体",1,2,3,4,5,6,7,8,'9-1','9-2',10,11,12]

#sheetから配列へ読み込み
def getSheet():
    global Respondents
    global Answer
    for i in range(2,MaxRow):
        team = sheet.cell(row=i,column=2).value
        if team == "9-1":
            team = 9
        elif team == '9-2':
            team = 10
        elif team >= 10:
            team = team + 1
            
        Respondents[team] = Respondents[team] + 1
        for j in range(3,15):
            q = j - 2
            cell = sheet.cell(row=i,column=j).value
            if cell == 1 : 
                Answer[team][q] = Answer[team][q] + 1

    for q in range(1,11):
        for team in range(1,14):
            Answer[0][q] = Answer[0][q] + Answer[team][q]

    pprint.pprint(Respondents)
    pprint.pprint(Answer)

def makePercentList():
    global Respondents
    global Answer
    global PercentNumber
    for team in range(1,14):
        Respondents[0] = Respondents[0] + Respondents[team]

    for team in range(0,14):
        for q in range(1,11):
#            PercentNumber[team][q] = int(Answer[team][q] / Respondents[team] *100)
            PercentNumber[team][q] = round(Answer[team][q] / Respondents[team] *100,1)

    pprint.pprint(PercentNumber)

def makeSheet4():
    global wb
    global Answer
    global Respondents
    global Label
    sheet = wb.create_sheet('Count')

    for team in range(0,14):
        sheet.cell(row=team + 2,column=1,value=Label[team])

    title  =["","夏祭り","福利厚生","防災","大掃除","ゴミ","防犯灯","交通標識","金杉会館","警察","渋滞","回答者数"]
#len(title) = 12
    for q in range(1,len(title)):
        sheet.cell(row=1,column=q+1,value=title[q])

    for team in range(0,14):
        for q in range(1,11):
            val = Answer[team][q]
            sheet.cell(row=team+2, column=q+1,value=val)
        sheet.cell(row=team+2, column=12,value=Respondents[team])
        

def makeSheet3():
    global PercentNumber
    global wb
    sheet3 = wb.create_sheet('Sheet3')

    for team in range(0,14):
        sheet3.cell(row=team + 2,column=1,value=Label[team])

    title  =["","夏祭り","福利厚生","防災","大掃除","ゴミ","防犯灯","交通標識","金杉会館","警察","渋滞"]
#    for q in range(1,11):
    for q in range(1,len(title))
        sheet3.cell(row=1,column=q+1,value=title[q])

    for team in range(0,14):
        max = 0
        min = 100
        for q in range(1,11):
            val = PercentNumber[team][q]
            sheet3.cell(row=team+2, column=q+1,value=val)
            if max <= val : max = val
            if min >= val : min = val
        for q in range(1,11):
            cell = sheet3.cell(row=team+2, column=q+1) 
            val = cell.value
            if max == val : cell.font = openpyxl.styles.fonts.Font(color='FF0000')
            if min == val : cell.font = openpyxl.styles.fonts.Font(color='0000FF')

        "=CORREL($B$3:$K$3,$B3:$K3)"

makeSheet2()
getSheet()
makePercentList()
makeSheet3()
makeSheet4()

wb.save('./tmp/アンケート結果3.xlsx')
